import os
import time
import threading
import requests
import datetime
import json
import io
import gzip
import uuid
import logging
from flask import Flask
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================== LOGGING ==================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================== CONFIG ==================
BOT_TOKEN = "8808046131:AAHCgB22O9KtwtIKrfXpMOBrPZRzNvN-3oo"
ADMIN_CHAT_ID = "2035024902"
CHANNEL_ID = "-1003903695158"

if not BOT_TOKEN or not ADMIN_CHAT_ID:
    raise RuntimeError("BOT_TOKEN and ADMIN_CHAT_ID must be set")

# ================== FILE PATHS ==================
MOTHER_FILE = "mother_accounts.json"
COOLDOWN_FILE = "user_cooldowns.json"
SUBSCRIBERS_FILE = "subscribers.json"
USER_INFO_FILE = "user_info.json"
ACCOUNTS_FILE = "accounts.json"
BALANCES_FILE = "balances.json"
DEPOSITS_FILE = "deposits.json"
CONFIG_FILE = "config.json"
SELL_REQUESTS_FILE = "sell_requests.json"
WITHDRAW_REQUESTS_FILE = "withdraw_requests.json"

app = Flask(__name__)

# ================== GLOBALS ==================
last_update_id = None
subscribed_users = set()
user_info = {}
mother_accounts = []
user_last_request = {}
submission_sessions = {}
support_sessions = set()
maintenance_mode = False

accounts = []
balances = {}
deposits = []
config = {
    "bkash_number": "",
    "price_per_account": 1.70,
    "sell_price_per_account": 1.0,
    "group_chat_id": "",
    "channel_id": str(CHANNEL_ID) if CHANNEL_ID else "",
    "maintenance_mode": False
}
deposit_sessions = {}
buy_sessions = set()
add_stock_sessions = {}
loss_recovery_sessions = {}
sell_sessions = {}
sell_requests = []
withdraw_requests = []
withdraw_sessions = {}
last_backup_message_id = None

data_lock = threading.RLock()
backup_lock = threading.Lock()

# ================== FILE I/O ==================
def load_mother_accounts():
    global mother_accounts
    try:
        with open(MOTHER_FILE, "r") as f:
            mother_accounts = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        mother_accounts = []

def save_mother_accounts():
    with data_lock:
        try:
            with open(MOTHER_FILE, "w") as f:
                json.dump(mother_accounts, f, indent=2)
        except IOError as e:
            logger.error(f"Mother save error: {e}")

def load_user_cooldowns():
    global user_last_request
    try:
        with open(COOLDOWN_FILE, "r") as f:
            user_last_request = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        user_last_request = {}

def save_user_cooldowns():
    with data_lock:
        try:
            with open(COOLDOWN_FILE, "w") as f:
                json.dump(user_last_request, f, indent=2)
        except IOError as e:
            logger.error(f"Cooldown save error: {e}")

def load_subscribers():
    global subscribed_users, user_info
    try:
        with open(SUBSCRIBERS_FILE, "r") as f:
            data = json.load(f)
            subscribed_users = set(data.get("subscribed", []))
    except (FileNotFoundError, json.JSONDecodeError):
        subscribed_users = set()
    try:
        with open(USER_INFO_FILE, "r") as f:
            user_info = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        user_info = {}

def save_subscribers():
    with data_lock:
        try:
            with open(SUBSCRIBERS_FILE, "w") as f:
                json.dump({"subscribed": list(subscribed_users)}, f)
        except IOError as e:
            logger.error(f"Subscribers save error: {e}")

def save_user_info():
    with data_lock:
        try:
            with open(USER_INFO_FILE, "w") as f:
                json.dump(user_info, f, indent=2)
        except IOError as e:
            logger.error(f"User info save error: {e}")

def load_market():
    global accounts, balances, deposits, config, CHANNEL_ID, maintenance_mode
    try:
        with open(ACCOUNTS_FILE, "r") as f:
            accounts = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        accounts = []
    try:
        with open(BALANCES_FILE, "r") as f:
            balances = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        balances = {}
    try:
        with open(DEPOSITS_FILE, "r") as f:
            deposits = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        deposits = []
    try:
        with open(CONFIG_FILE, "r") as f:
            loaded_config = json.load(f)
            for key, value in config.items():
                if key not in loaded_config:
                    loaded_config[key] = value
            config = loaded_config
            CHANNEL_ID = int(config.get("channel_id", "0")) if config.get("channel_id") else None
            maintenance_mode = config.get("maintenance_mode", False)
    except (FileNotFoundError, json.JSONDecodeError):
        config["channel_id"] = str(CHANNEL_ID) if CHANNEL_ID else ""
        maintenance_mode = False
        CHANNEL_ID = int(config["channel_id"]) if config["channel_id"] else None

def save_accounts():
    with data_lock:
        try:
            with open(ACCOUNTS_FILE, "w") as f:
                json.dump(accounts, f, indent=2)
        except IOError as e:
            logger.error(f"Accounts save error: {e}")

def save_balances():
    with data_lock:
        try:
            with open(BALANCES_FILE, "w") as f:
                json.dump(balances, f, indent=2)
        except IOError as e:
            logger.error(f"Balances save error: {e}")

def save_deposits():
    with data_lock:
        try:
            with open(DEPOSITS_FILE, "w") as f:
                json.dump(deposits, f, indent=2)
        except IOError as e:
            logger.error(f"Deposits save error: {e}")

def save_config():
    with data_lock:
        config["maintenance_mode"] = maintenance_mode
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(config, f, indent=2)
        except IOError as e:
            logger.error(f"Config save error: {e}")

def load_sell_requests():
    global sell_requests
    try:
        with open(SELL_REQUESTS_FILE, "r") as f:
            sell_requests = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        sell_requests = []

def save_sell_requests():
    with data_lock:
        try:
            with open(SELL_REQUESTS_FILE, "w") as f:
                json.dump(sell_requests, f, indent=2)
        except IOError as e:
            logger.error(f"Sell requests save error: {e}")

def load_withdraw_requests():
    global withdraw_requests
    try:
        with open(WITHDRAW_REQUESTS_FILE, "r") as f:
            withdraw_requests = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        withdraw_requests = []

def save_withdraw_requests():
    with data_lock:
        try:
            with open(WITHDRAW_REQUESTS_FILE, "w") as f:
                json.dump(withdraw_requests, f, indent=2)
        except IOError as e:
            logger.error(f"Withdraw requests save error: {e}")

def save_all():
    save_accounts()
    save_balances()
    save_deposits()
    save_config()
    save_subscribers()
    save_user_info()
    save_mother_accounts()
    save_user_cooldowns()
    save_sell_requests()
    save_withdraw_requests()
    save_data_to_channel()

# ================== TELEGRAM HELPERS ==================
def send_telegram_message(text, chat_id, reply_markup=None, parse_mode=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    if parse_mode:
        payload["parse_mode"] = parse_mode
    for attempt in range(3):
        try:
            resp = requests.post(url, json=payload, timeout=10)
            if resp.status_code == 429:
                retry_after = resp.json().get("parameters", {}).get("retry_after", 2)
                time.sleep(retry_after)
                continue
            return resp
        except Exception as e:
            logger.error(f"Send error to {chat_id} (attempt {attempt+1}): {e}")
            if attempt == 2:
                return None
            time.sleep(1)
    return None

def send_telegram_document(file_bytes, filename, chat_id, caption=""):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    try:
        files = {'document': (filename, file_bytes,
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = requests.post(url, data={"chat_id": chat_id, "caption": caption}, files=files, timeout=30)
        return resp.status_code == 200 and resp.json().get("ok", False)
    except Exception as e:
        logger.error(f"Document send error: {e}")
        return False

def delete_telegram_message(chat_id, message_id):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage"
    try:
        requests.post(url, json={"chat_id": chat_id, "message_id": message_id}, timeout=10)
    except Exception as e:
        logger.error(f"Delete message error: {e}")

def forward_telegram_document(chat_id, from_chat_id, message_id):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/forwardMessage"
    payload = {"chat_id": chat_id, "from_chat_id": from_chat_id, "message_id": message_id}
    try:
        requests.post(url, json=payload, timeout=10)
    except Exception as e:
        logger.error(f"Forward error: {e}")

def broadcast_message(text):
    to_remove = []
    for chat_id in list(subscribed_users):
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        try:
            resp = requests.post(url, json={"chat_id": chat_id, "text": text}, timeout=5)
            if resp.status_code == 403:
                to_remove.append(chat_id)
            elif resp.status_code == 429:
                retry_after = resp.json().get("parameters", {}).get("retry_after", 1)
                time.sleep(retry_after)
                requests.post(url, json={"chat_id": chat_id, "text": text}, timeout=5)
        except Exception as e:
            logger.error(f"Broadcast to {chat_id} failed: {e}")
            to_remove.append(chat_id)
        time.sleep(0.05)
    if to_remove:
        with data_lock:
            for uid in to_remove:
                subscribed_users.discard(uid)
                user_info.pop(uid, None)
        save_subscribers()
        save_user_info()

# ================== CHANNEL BACKUP ==================
def save_data_to_channel():
    global last_backup_message_id
    if not CHANNEL_ID:
        return
    with backup_lock:
        try:
            with data_lock:
                data = {
                    "accounts": accounts,
                    "balances": balances,
                    "deposits": deposits,
                    "config": config,
                    "subscribed_users": list(subscribed_users),
                    "user_info": user_info,
                    "mother_accounts": mother_accounts,
                    "user_cooldowns": user_last_request,
                    "sell_requests": sell_requests,
                    "withdraw_requests": withdraw_requests,
                    "timestamp": datetime.datetime.now().isoformat()
                }
            json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode('utf-8')
            compressed = gzip.compress(json_bytes, compresslevel=6)
            max_size = 48 * 1024 * 1024
            if len(compressed) > max_size:
                logger.warning("Compressed backup too large, cannot upload to Telegram")
                return

            if last_backup_message_id:
                try:
                    delete_telegram_message(CHANNEL_ID, last_backup_message_id)
                except:
                    pass

            filename = f"backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json.gz"
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
            files = {'document': (filename, compressed, 'application/gzip')}
            resp = requests.post(url, data={"chat_id": CHANNEL_ID}, files=files, timeout=60)

            if resp.status_code == 200 and resp.json().get("ok"):
                last_backup_message_id = resp.json()["result"]["message_id"]
                pin_url = f"https://api.telegram.org/bot{BOT_TOKEN}/pinChatMessage"
                requests.post(pin_url, json={
                    "chat_id": CHANNEL_ID,
                    "message_id": last_backup_message_id,
                    "disable_notification": True
                })
            else:
                logger.error(f"Backup upload failed: {resp.text}")

        except Exception as e:
            logger.error(f"Channel backup error: {e}")

def auto_backup_loop():
    while True:
        time.sleep(86400)
        save_data_to_channel()
        send_telegram_message("🔄 অটো ব্যাকআপ সম্পন্ন হয়েছে", ADMIN_CHAT_ID)

# ================== KEYBOARD ==================
def get_keyboard(chat_id):
    keyboard = [
        ["💰 ব্যালেন্স", "💸 ডিপোজিট"],
        ["🛒 একাউন্ট কিনুন", "💰 বিক্রয় করুন"],
        ["💸 উইথড্র"],
        ["📋 সাবমিট", "🎁 মাদার একাউন্ট"],
        ["📞 সাপোর্ট", "🛑 স্টপ"],
        ["🔄 লস রিকভারি"]
    ]
    if str(chat_id) == ADMIN_CHAT_ID:
        keyboard.append(["📥 ডিপোজিট রিকোয়েস্ট", "➕ স্টক যোগ করুন"])
        keyboard.append(["📦 স্টক দেখুন", "🗑️ স্টক ডিলিট"])
        keyboard.append(["📊 সেল requests", "💳 withdraw requests"])
    return {"keyboard": keyboard, "resize_keyboard": True, "one_time_keyboard": False}

def remove_keyboard():
    return {"remove_keyboard": True}

def send_main_keyboard(chat_id, text="\u200b"):
    send_telegram_message(text, chat_id, reply_markup=get_keyboard(chat_id))

# ================== EXCEL GENERATORS ==================
def generate_submission_excel(usernames, passwords, twofa_list, bkash, telegram_username):
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Submission"
    headers = ["Username", "Password", "2FA Key", "Bkash Number", "Telegram Username"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i in range(len(usernames)):
        row = [
            usernames[i],
            passwords[i] if i < len(passwords) else "",
            twofa_list[i] if i < len(twofa_list) else "",
            bkash,
            telegram_username
        ]
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def generate_purchase_excel(bought):
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchased Accounts"
    headers = ["Username", "Password", "2FA Key"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for acc in bought:
        ws.append([acc["username"], acc["password"], acc.get("fa_key", "")])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def generate_sell_excel(accounts_list, telegram_username):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sell Request"
    headers = ["Username", "Password", "2FA Key", "Telegram Username"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for acc in accounts_list:
        ws.append([acc["username"], acc["password"], acc.get("fa_key", ""), telegram_username])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

# ================== SUBMISSION HANDLER ==================
def start_submission(chat_id, sender_username):
    submission_sessions[chat_id] = {"step": "username", "data": {}, "username": sender_username}
    send_telegram_message(
        "📋 দয়া করে আপনার **ইউজারনেম** লিস্ট দিন (প্রতি লাইনে একটি করে):\n\n"
        "উদাহরণ:\nuser1\nuser2\nuser3\n\n/start দিয়ে আবার শুরু করতে পারেন।",
        chat_id
    )

def process_submission_step(chat_id, text, sender_username):
    if chat_id not in submission_sessions:
        return False
    if text.strip().lower() == "/start":
        del submission_sessions[chat_id]
        send_telegram_message("❌ জমা প্রক্রিয়া বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = submission_sessions[chat_id]
    step = session["step"]
    if step == "username":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id)
            return True
        session["data"]["usernames"] = lines
        session["step"] = "password"
        send_telegram_message(f"🔑 এখন **পাসওয়ার্ড** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nআপনার ইউজারনেম সংখ্যা: {len(lines)}", chat_id)
        return True
    elif step == "password":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        usernames = session["data"]["usernames"]
        if len(lines) != len(usernames):
            send_telegram_message(f"❌ পাসওয়ার্ড সংখ্যা ({len(lines)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। পুনরায় সঠিক লিস্ট দিন।", chat_id)
            return True
        session["data"]["passwords"] = lines
        session["step"] = "2fa"
        send_telegram_message("🔐 এখন **2FA কী** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nযদি 2FA না থাকে, লাইন ফাঁকা রাখবেন (শুধু এন্টার দিন)।", chat_id)
        return True
    elif step == "2fa":
        raw_lines = text.splitlines()
        twofa_list = [l.strip() for l in raw_lines]
        usernames = session["data"]["usernames"]
        while len(twofa_list) > len(usernames) and twofa_list and twofa_list[-1] == '':
            twofa_list.pop()
        if len(twofa_list) != len(usernames):
            send_telegram_message(f"❌ 2FA কী সংখ্যা ({len(twofa_list)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। প্রতিটি ইউজারনেমের জন্য একটি লাইন (খালিও হতে পারে) দিন।", chat_id)
            return True
        session["data"]["twofa"] = twofa_list
        session["step"] = "bkash"
        send_telegram_message("💳 দয়া করে আপনার **বিকাশ নম্বর** দিন:", chat_id)
        return True
    elif step == "bkash":
        bkash_number = text.strip()
        if not bkash_number:
            send_telegram_message("⚠️ বিকাশ নম্বর খালি রাখা যাবে না। আবার দিন।", chat_id)
            return True
        session["data"]["bkash"] = bkash_number
        excel_bytes = generate_submission_excel(
            session["data"]["usernames"], session["data"]["passwords"],
            session["data"]["twofa"], bkash_number, session["username"]
        )
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"submission_{chat_id}_{timestamp}.xlsx"
        if send_telegram_document(excel_bytes, filename, ADMIN_CHAT_ID):
            send_telegram_message("✅ আপনার অ্যাকাউন্ট সফলভাবে জমা হয়েছে।\n\nঅ্যাডমিন শীঘ্রই যোগাযোগ করবে। ধন্যবাদ! 🙏", chat_id)
        else:
            send_telegram_message("⚠️ জমা দেওয়ার সময় ত্রুটি হয়েছে, অনুগ্রহ করে পরে চেষ্টা করুন।", chat_id)
        del submission_sessions[chat_id]
        send_main_keyboard(chat_id)
        return True
    return False

# ================== LOSS RECOVERY HANDLER ==================
def start_loss_recovery(chat_id):
    loss_recovery_sessions[chat_id] = {"step": "usernames", "data": {}}
    send_telegram_message(
        "⚠️ সতর্কতা: ভুল তথ্য দিলে লস রিকভারি পাবেন না। সকল তথ্য ম্যানুয়ালি যাচাই করা হবে।\n\n"
        "অনুগ্রহ করে সঠিক তথ্য দিন।\n\n"
        "আপনার কেনা অ্যাকাউন্টগুলোর ইউজারনেম লিস্ট দিন (প্রতি লাইনে একটি):",
        chat_id
    )

def process_loss_recovery_step(chat_id, text):
    if chat_id not in loss_recovery_sessions:
        return False
    if text.strip().lower() == "/start":
        del loss_recovery_sessions[chat_id]
        send_telegram_message("❌ লস রিকভারি বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = loss_recovery_sessions[chat_id]
    step = session["step"]
    if step == "usernames":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id)
            return True
        session["data"]["usernames"] = lines
        session["step"] = "cookie_date"
        send_telegram_message("📅 কত তারিখে কুকিজ সাবমিট করেছিলেন? (শুধু তারিখের সংখ্যা লিখুন, যেমন: 13 বা 26)", chat_id)
        return True
    elif step == "cookie_date":
        date_str = text.strip()
        if not date_str.isdigit():
            send_telegram_message("⚠️ দয়া করে শুধু সংখ্যা দিন (13, 26 ইত্যাদি)।", chat_id)
            return True
        session["data"]["cookie_date"] = date_str
        session["step"] = "report_file"
        send_telegram_message("📎 এখন রিপোর্ট ফেইল হওয়ার দিনের Excel Report File (.xlsx/.xls) পাঠান।\n\n⚠️ Screenshot, PDF বা অন্য কোনো ফাইল গ্রহণ করা হবে না।", chat_id)
        return True
    elif step == "report_file":
        send_telegram_message("⚠️ শুধুমাত্র Excel File (.xlsx/.xls) পাঠান।", chat_id)
        return True
    elif step == "bkash":
        bkash = text.strip()
        if not bkash:
            send_telegram_message("⚠️ বিকাশ নম্বর খালি রাখা যাবে না।", chat_id)
            return True
        session["data"]["bkash"] = bkash
        session["step"] = "whatsapp"
        send_telegram_message("📞 আপনার হোয়াটসঅ্যাপ নম্বর দিন:", chat_id)
        return True
    elif step == "whatsapp":
        whatsapp = text.strip()
        if not whatsapp:
            send_telegram_message("⚠️ হোয়াটসঅ্যাপ নম্বর দিন।", chat_id)
            return True
        session["data"]["whatsapp"] = whatsapp
        usernames = session["data"]["usernames"]
        cookie_date = session["data"]["cookie_date"]
        bkash = session["data"]["bkash"]
        file_id = session["data"].get("report_file_id")
        file_message_id = session["data"].get("report_message_id")
        admin_text = (
            "🔄 **নতুন লস রিকভারি রিকোয়েস্ট**\n\n"
            f"👤 ইউজার: {user_info.get(chat_id, chat_id)} (`{chat_id}`)\n"
            f"📅 কুকি সাবমিটের তারিখ: {cookie_date}\n"
            f"💳 বিকাশ: {bkash}\n"
            f"📞 হোয়াটসঅ্যাপ: {whatsapp}\n"
            f"🔑 ইউজারনেম: " + ", ".join(usernames)
        )
        send_telegram_message(admin_text, ADMIN_CHAT_ID, parse_mode="Markdown")
        if file_message_id:
            forward_telegram_document(ADMIN_CHAT_ID, chat_id, file_message_id)
        elif file_id:
            send_telegram_message("⚠️ রিপোর্ট ফাইল ফরওয়ার্ড করা যায়নি, কারণ মেসেজ আইডি পাওয়া যায়নি।", ADMIN_CHAT_ID)
        send_telegram_message("✅ আপনার লস রিকভারি রিকোয়েস্ট জমা হয়েছে। অ্যাডমিন শীঘ্রই আপনার সাথে যোগাযোগ করবে।", chat_id)
        del loss_recovery_sessions[chat_id]
        send_main_keyboard(chat_id)
        return True
    return False

def handle_loss_recovery_file(chat_id, message):
    if chat_id not in loss_recovery_sessions:
        return
    session = loss_recovery_sessions[chat_id]
    if session["step"] != "report_file":
        return
    doc = message.get("document")
    if not doc:
        send_telegram_message("⚠️ শুধুমাত্র Excel File (.xlsx/.xls) পাঠান।", chat_id)
        return
    file_name = doc.get("file_name", "")
    if not file_name.lower().endswith(('.xlsx', '.xls')):
        send_telegram_message("❌ শুধুমাত্র .xlsx বা .xls এক্সটেনশনের ফাইল গ্রহণ করা হবে।", chat_id)
        return
    allowed_mimes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    ]
    mime = doc.get("mime_type", "")
    if mime not in allowed_mimes:
        send_telegram_message("❌ অবৈধ ফাইল ফরম্যাট। শুধুমাত্র এক্সেল ফাইল গ্রহণ করা হবে।", chat_id)
        return
    file_id = doc.get("file_id")
    message_id = message.get("message_id")
    session["data"]["report_file_id"] = file_id
    session["data"]["report_message_id"] = message_id
    session["step"] = "bkash"
    send_telegram_message("💳 দয়া করে আপনার বিকাশ নম্বর দিন (যেটি ব্যবহার করেছিলেন):", chat_id)

# ================== MOTHER ACCOUNT ==================
def handle_addmother(chat_id, args):
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    parts = args.split(maxsplit=2)
    if len(parts) < 2:
        send_telegram_message("❌ ফরম্যাট: /addmother username password [2fa_key]", chat_id)
        return
    username = parts[0]
    password = parts[1]
    fa_key = parts[2] if len(parts) == 3 else ""
    with data_lock:
        for acc in mother_accounts:
            if acc["username"] == username and acc["password"] == password:
                send_telegram_message("⚠️ এই অ্যাকাউন্টটি আগেই যোগ করা আছে।", chat_id)
                return
        mother_accounts.append({"username": username, "password": password, "fa_key": fa_key, "assigned_to": None, "assigned_at": None})
        save_mother_accounts()
    save_data_to_channel()
    send_telegram_message(f"✅ মাদার অ্যাকাউন্ট যোগ করা হয়েছে: {username}", chat_id)

def handle_getmother(chat_id):
    now = time.time()
    last = user_last_request.get(str(chat_id), 0)
    cooldown = 600
    if now - last < cooldown:
        wait_sec = cooldown - (now - last)
        wait_min = int(wait_sec // 60)
        wait_sec_rem = int(wait_sec % 60)
        send_telegram_message(f"⏳ অনুগ্রহ করে অপেক্ষা করুন। পরবর্তী অ্যাকাউন্ট {wait_min} মিনিট {wait_sec_rem} সেকেন্ড পর নিতে পারবেন।", chat_id)
        send_main_keyboard(chat_id)
        return
    selected_acc = None
    with data_lock:
        for acc in mother_accounts:
            if acc["assigned_to"] is None:
                selected_acc = acc
                break
        if selected_acc is None:
            send_telegram_message("❌ কোনো মাদার অ্যাকাউন্ট উপলব্ধ নেই। পরে আবার চেষ্টা করুন।", chat_id)
            send_main_keyboard(chat_id)
            return
    msg = f"🎁 আপনার মাদার অ্যাকাউন্ট:\n\n👤 ইউজারনেম: {selected_acc['username']}\n🔑 পাসওয়ার্ড: {selected_acc['password']}"
    if selected_acc["fa_key"]:
        msg += f"\n🔐 2FA Key: {selected_acc['fa_key']}"
    if send_telegram_message(msg, chat_id):
        with data_lock:
            selected_acc["assigned_to"] = str(chat_id)
            selected_acc["assigned_at"] = now
            user_last_request[str(chat_id)] = now
            save_mother_accounts()
            save_user_cooldowns()
        send_main_keyboard(chat_id)
    else:
        send_telegram_message("⚠️ মেসেজ পাঠানো যায়নি, পরে আবার চেষ্টা করুন।", chat_id)
    save_data_to_channel()

def handle_motherlist(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    with data_lock:
        if not mother_accounts:
            send_telegram_message("📭 কোনো মাদার অ্যাকাউন্ট নেই।", chat_id)
            return
        lines = ["🎁 *মাদার অ্যাকাউন্ট লিস্ট:*\n"]
        for i, acc in enumerate(mother_accounts, start=1):
            assigned = "কেহ না"
            if acc["assigned_to"]:
                try:
                    assigned_time = datetime.datetime.fromtimestamp(acc["assigned_at"]).strftime('%d/%m %H:%M')
                except:
                    assigned_time = "কিছুক্ষণ আগে"
                assigned = f"{acc['assigned_to']} ({assigned_time})"
            twofa = "আছে" if acc.get("fa_key") else "নেই"
            lines.append(f"{i}. ইউজার: {acc['username']} | পাস: {acc['password']} | 2FA: {twofa} | বরাদ্দ: {assigned}")
    send_telegram_message("\n".join(lines), chat_id)

def handle_deletemother(chat_id, arg):
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    arg = arg.strip()
    with data_lock:
        try:
            idx = int(arg) - 1
            if 0 <= idx < len(mother_accounts):
                deleted = mother_accounts.pop(idx)
                save_mother_accounts()
            else:
                send_telegram_message("❌ ভুল ইনডেক্স। /motherlist দিয়ে নম্বর দেখুন।", chat_id)
                return
        except ValueError:
            for i, acc in enumerate(mother_accounts):
                if acc["username"] == arg:
                    deleted = mother_accounts.pop(i)
                    save_mother_accounts()
                    break
            else:
                send_telegram_message(f"❌ `{arg}` নামে কোনো মাদার অ্যাকাউন্ট পাওয়া যায়নি।", chat_id)
                return
    save_data_to_channel()
    send_telegram_message(f"✅ মাদার অ্যাকাউন্ট `{deleted['username']}` মুছে ফেলা হয়েছে।", chat_id)

# ================== MAINTENANCE MODE ==================
def handle_maintenance(chat_id, args):
    global maintenance_mode
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    args = args.strip().lower()
    if args == "on":
        maintenance_mode = True
        save_config()
        send_telegram_message("🔧 রক্ষণাবেক্ষণ মোড চালু করা হয়েছে। সাধারণ ইউজাররা এখন বট ব্যবহার করতে পারবে না।", chat_id)
    elif args == "off":
        maintenance_mode = False
        save_config()
        send_telegram_message("🔧 রক্ষণাবেক্ষণ মোড বন্ধ করা হয়েছে। বট এখন স্বাভাবিক ভাবে চলবে।", chat_id)
    else:
        status = "চালু" if maintenance_mode else "বন্ধ"
        send_telegram_message(f"🔧 রক্ষণাবেক্ষণ মোড বর্তমানে {status} আছে। /maintenance on/off দিয়ে পরিবর্তন করুন।", chat_id)

# ================== ADMIN BROADCAST & USERS ==================
def handle_admin_users(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return
    with data_lock:
        if not subscribed_users:
            send_telegram_message("কোনো সাবস্ক্রাইবার নেই।", chat_id)
            return
        msg_lines = ["📋 সাবস্ক্রাইবড ইউজার লিস্ট:\n"]
        for uid in subscribed_users:
            name = user_info.get(str(uid), f"ID:{uid}")
            if ' ' not in name:
                name = '@' + name
            msg_lines.append(f"• {name} (ID: {uid})")
    send_telegram_message("\n".join(msg_lines), chat_id)

def handle_admin_broadcast(chat_id, message):
    if str(chat_id) != ADMIN_CHAT_ID:
        return
    if not message.strip():
        send_telegram_message("❌ মেসেজ খালি রাখা যাবে না। ফরম্যাট: /broadcast <মেসেজ>", chat_id)
        return
    broadcast_message(f"📢 অ্যাডমিন থেকে বার্তা:\n\n{message}")
    send_telegram_message("✅ বার্তা সকল সাবস্ক্রাইবারকে পাঠানো হয়েছে।", chat_id)

def handle_admin_send(chat_id, target_id, message):
    if str(chat_id) != ADMIN_CHAT_ID:
        return
    if not target_id.isdigit():
        send_telegram_message("❌ সঠিক ইউজার আইডি দিন।", chat_id)
        return
    if not message.strip():
        send_telegram_message("❌ মেসেজ খালি রাখা যাবে না।", chat_id)
        return
    send_telegram_message(f"📩 অ্যাডমিন থেকে:\n\n{message}", target_id)
    send_telegram_message(f"✅ {target_id} কে মেসেজ পাঠানো হয়েছে।", chat_id)

# ================== BACKUP & RESTORE ==================
def handle_backup(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    save_all()
    with data_lock:
        backup = {
            "subscribed_users": list(subscribed_users),
            "user_info": user_info,
            "mother_accounts": mother_accounts,
            "user_cooldowns": user_last_request,
            "accounts": accounts,
            "balances": balances,
            "deposits": deposits,
            "config": config,
            "sell_requests": sell_requests,
            "withdraw_requests": withdraw_requests,
            "timestamp": datetime.datetime.now().isoformat()
        }
    backup_json = json.dumps(backup, indent=2, ensure_ascii=False).encode('utf-8')
    compressed = gzip.compress(backup_json)
    filename = f"backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json.gz"
    if send_telegram_document(compressed, filename, ADMIN_CHAT_ID, caption="Manual backup"):
        send_telegram_message("✅ ব্যাকআপ ফাইল তৈরি ও পাঠানো হয়েছে।", chat_id)
    else:
        send_telegram_message("⚠️ ব্যাকআপ ফাইল পাঠানো যায়নি।", chat_id)

def handle_restore(chat_id, file_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        send_telegram_message("❌ আপনি এই কমান্ড ব্যবহার করতে পারবেন না।", chat_id)
        return
    if _perform_restore(file_id):
        send_telegram_message("✅ ব্যাকআপ রিস্টোর সম্পন্ন হয়েছে।", chat_id)
    else:
        send_telegram_message("❌ ব্যাকআপ রিস্টোর করতে ব্যর্থ হয়েছে।", chat_id)

def _perform_restore(file_id):
    try:
        get_file_url = f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}"
        resp = requests.get(get_file_url, timeout=20)
        resp.raise_for_status()
        file_data = resp.json()
        if not file_data.get("ok"):
            logger.error("getFile failed")
            return False
        file_path = file_data["result"]["file_path"]
        download_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"
        file_resp = requests.get(download_url, timeout=60)
        file_resp.raise_for_status()
        content = file_resp.content
        try:
            content = gzip.decompress(content)
        except gzip.BadGzipFile:
            pass
        backup = json.loads(content.decode('utf-8'))
    except Exception as e:
        logger.error(f"_perform_restore error: {e}")
        return False

    with data_lock:
        global subscribed_users, user_info, mother_accounts, user_last_request
        global accounts, balances, deposits, config, CHANNEL_ID, maintenance_mode
        global sell_requests, withdraw_requests
        subscribed_users = set(backup.get("subscribed_users", []))
        user_info = backup.get("user_info", {})
        mother_accounts = backup.get("mother_accounts", [])
        user_last_request = backup.get("user_cooldowns", {})
        accounts = backup.get("market_accounts", backup.get("accounts", []))
        balances = backup.get("balances", {})
        deposits = backup.get("deposits", [])
        config = backup.get("config", {})
        CHANNEL_ID = int(config.get("channel_id", "0")) if config.get("channel_id") else None
        maintenance_mode = config.get("maintenance_mode", False)
        sell_requests = backup.get("sell_requests", [])
        withdraw_requests = backup.get("withdraw_requests", [])
        save_all()
    logger.info("Restore completed successfully")
    return True

def auto_restore_from_channel():
    if not CHANNEL_ID:
        logger.info("Auto-restore skipped: No CHANNEL_ID")
        return
    try:
        get_chat_url = f"https://api.telegram.org/bot{BOT_TOKEN}/getChat?chat_id={CHANNEL_ID}"
        resp = requests.get(get_chat_url, timeout=20).json()
        if not resp.get("ok"):
            logger.warning("Could not get channel info for auto-restore")
            return
        pinned = resp["result"].get("pinned_message")
        if not pinned or "document" not in pinned:
            logger.info("No pinned document in channel; skipping auto-restore")
            return
        file_id = pinned["document"]["file_id"]
        logger.info("Auto-restore: Found pinned backup. Restoring...")
        if _perform_restore(file_id):
            logger.info("Auto-restore completed successfully")
        else:
            logger.error("Auto-restore failed")
    except Exception as e:
        logger.error(f"Auto-restore error: {e}")

# ================== ADMIN ADD STOCK FLOW ==================
def start_add_stock(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return
    add_stock_sessions[chat_id] = {"step": "usernames"}
    send_telegram_message(
        "➕ স্টক যোগ করুন\n\nপ্রথমে **ইউজারনেম** লিস্ট দিন (প্রতি লাইনে একটি করে):\nউদাহরণ:\nuser1\nuser2\nuser3\n\n/start দিয়ে বাতিল করুন।",
        chat_id
    )

def process_add_stock_step(chat_id, text):
    if chat_id not in add_stock_sessions:
        return False
    if text.strip().lower() == "/start":
        del add_stock_sessions[chat_id]
        send_telegram_message("❌ স্টক যোগ বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = add_stock_sessions[chat_id]
    step = session["step"]
    if step == "usernames":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id)
            return True
        session["usernames"] = lines
        session["step"] = "passwords"
        send_telegram_message(f"🔑 এখন **পাসওয়ার্ড** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nআপনার ইউজারনেম সংখ্যা: {len(lines)}", chat_id)
        return True
    elif step == "passwords":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        usernames = session["usernames"]
        if len(lines) != len(usernames):
            send_telegram_message(f"❌ পাসওয়ার্ড সংখ্যা ({len(lines)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। পুনরায় সঠিক লিস্ট দিন।", chat_id)
            return True
        session["passwords"] = lines
        session["step"] = "fa_keys"
        send_telegram_message("🔐 এখন **2FA কী** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nযদি 2FA না থাকে, লাইন ফাঁকা রাখবেন (শুধু এন্টার দিন)।", chat_id)
        return True
    elif step == "fa_keys":
        raw_lines = text.splitlines()
        fa_list = [l.strip() for l in raw_lines]
        usernames = session["usernames"]
        while len(fa_list) > len(usernames) and fa_list and fa_list[-1] == '':
            fa_list.pop()
        if len(fa_list) != len(usernames):
            send_telegram_message(f"❌ 2FA কী সংখ্যা ({len(fa_list)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। প্রতিটি ইউজারনেমের জন্য একটি লাইন (খালিও হতে পারে) দিন।", chat_id)
            return True
        count = len(usernames)
        with data_lock:
            for i in range(count):
                accounts.append({"username": usernames[i], "password": session["passwords"][i], "fa_key": fa_list[i]})
            save_accounts()
        save_data_to_channel()
        del add_stock_sessions[chat_id]
        send_telegram_message(f"✅ {count} টি অ্যাকাউন্ট স্টকে যোগ করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    return False

# ================== MARKETPLACE: DEPOSIT & BUY ==================
def start_deposit(chat_id):
    deposit_sessions[chat_id] = {"step": "amount"}
    bkash = config.get("bkash_number", "")
    if not bkash:
        send_telegram_message("⚠️ অ্যাডমিন এখনও বিকাশ নম্বর সেট করেননি। পরে চেষ্টা করুন।", chat_id)
        send_main_keyboard(chat_id)
        deposit_sessions.pop(chat_id, None)
        return
    send_telegram_message(f"💸 দয়া করে আপনার জমা করার টাকার পরিমাণ লিখুন (শুধু সংখ্যা, যেমন: 100)\n\nবিকাশ নম্বর: {bkash}\n\nটাকা পাঠানোর পর ট্রানজেকশন আইডি সহ পুনরায় লিখবেন।", chat_id)

def process_deposit_step(chat_id, text):
    if chat_id not in deposit_sessions:
        return False
    if text.strip().lower() == "/start":
        deposit_sessions.pop(chat_id, None)
        send_telegram_message("❌ ডিপোজিট বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = deposit_sessions[chat_id]
    step = session["step"]
    if step == "amount":
        try:
            amount = float(text)
            if amount <= 0:
                raise ValueError
        except:
            send_telegram_message("⚠️ সঠিক সংখ্যা দিন। /start দিয়ে বাতিল করুন।", chat_id)
            return True
        session["amount"] = amount
        session["step"] = "trxid"
        send_telegram_message("🔢 এখন আপনার বিকাশ ট্রানজেকশন আইডি লিখুন:", chat_id)
        return True
    elif step == "trxid":
        trxid = text.strip()
        if not trxid:
            send_telegram_message("⚠️ ট্রানজেকশন আইডি খালি রাখা যাবে না।", chat_id)
            return True
        amount = session["amount"]
        deposit_id = uuid.uuid4().hex[:10]
        deposit = {"id": deposit_id, "user_id": chat_id, "amount": amount, "trxid": trxid, "status": "pending", "time": time.time()}
        with data_lock:
            deposits.append(deposit)
            save_deposits()
        save_data_to_channel()
        deposit_sessions.pop(chat_id, None)
        admin_msg = f"📥 নতুন ডিপোজিট রিকোয়েস্ট\nআইডি: {deposit_id}\nইউজার: {user_info.get(chat_id, chat_id)} ({chat_id})\nপরিমাণ: {amount} টাকা\nট্রানজেকশন আইডি: {trxid}\nঅনুমোদন করতে: /approve {deposit_id}\nবাতিল করতে: /reject {deposit_id}"
        send_telegram_message(admin_msg, ADMIN_CHAT_ID)
        send_telegram_message(f"✅ আপনার {amount} টাকার ডিপোজিট রিকোয়েস্ট জমা হয়েছে।\nঅ্যাডমিন অনুমোদন করলেই আপনার ব্যালেন্সে যোগ হবে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    return False

def handle_buy(chat_id, quantity):
    try:
        qty = int(quantity)
        if qty <= 0:
            raise ValueError
    except:
        send_telegram_message("❌ সঠিক সংখ্যা দিন। যেমন: 3", chat_id)
        return False
    with data_lock:
        if qty > len(accounts):
            send_telegram_message(f"❌ পর্যাপ্ত অ্যাকাউন্ট নেই। বর্তমান স্টক: {len(accounts)}", chat_id)
            return False
        price = config.get("price_per_account", 1.70)
        total = qty * price
        user_balance = balances.get(str(chat_id), 0)
        if user_balance < total:
            send_telegram_message(f"❌ পর্যাপ্ত ব্যালেন্স নেই।\nপ্রয়োজন: {total} টাকা\nআপনার ব্যালেন্স: {user_balance} টাকা\nদয়া করে প্রথমে ডিপোজিট করুন।", chat_id)
            return False
        bought = accounts[:qty]
        del accounts[:qty]
        balances[str(chat_id)] = user_balance - total
        save_accounts()
        save_balances()
    excel_bytes = generate_purchase_excel(bought)
    filename = f"purchased_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if send_telegram_document(excel_bytes, filename, chat_id):
        send_telegram_message(f"✅ {qty} টি অ্যাকাউন্ট কেনা হয়েছে। মোট মূল্য: {total} টাকা।\nঅবশিষ্ট ব্যালেন্স: {balances[str(chat_id)]} টাকা", chat_id)
        admin_msg = f"🛒 {user_info.get(chat_id, chat_id)} ({chat_id}) {qty} টি অ্যাকাউন্ট কিনেছে। মোট: {total} টাকা।"
        send_telegram_message(admin_msg, ADMIN_CHAT_ID)
    else:
        with data_lock:
            accounts[:0] = bought
            balances[str(chat_id)] = user_balance
            save_accounts()
            save_balances()
        send_telegram_message("⚠️ অ্যাকাউন্ট ডেলিভারি ব্যর্থ হয়েছে। আপনার টাকা ফেরত দেওয়া হয়েছে এবং অ্যাকাউন্ট পুনরায় স্টকে যোগ করা হয়েছে। পরে আবার চেষ্টা করুন।", chat_id)
    save_data_to_channel()
    return True

# ================== SELL ==================
def start_sell(chat_id):
    send_telegram_message("⚠️ সতর্কতা:\n\nআপনি যে একাউন্ট গুলো সেল দিবেন সেগুলো আমাদের ইউজাররা কুকিজ সাবমিট করে যদি তাদের লস হয় যে পরিমাণ টাকা লস হবে তা আপনার সেল এমাউন্ট হতে মাইনাস হবে।", chat_id)
    sell_price = config.get("sell_price_per_account", 1.0)
    sell_sessions[chat_id] = {"step": "usernames"}
    send_telegram_message(
        f"💰 বিক্রয় করুন\n\nপ্রতি অ্যাকাউন্টের মূল্য: {sell_price} টাকা\n\nপ্রথমে আপনার **ইউজারনেম** লিস্ট দিন (প্রতি লাইনে একটি করে):\nউদাহরণ:\nuser1\nuser2\nuser3\n\n/start দিয়ে বাতিল করুন।",
        chat_id
    )

def process_sell_step(chat_id, text):
    if chat_id not in sell_sessions:
        return False
    if text.strip().lower() == "/start":
        del sell_sessions[chat_id]
        send_telegram_message("❌ বিক্রয় প্রক্রিয়া বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = sell_sessions[chat_id]
    step = session["step"]
    if step == "usernames":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id)
            return True
        session["usernames"] = lines
        session["step"] = "passwords"
        send_telegram_message(f"🔑 এখন **পাসওয়ার্ড** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nআপনার ইউজারনেম সংখ্যা: {len(lines)}", chat_id)
        return True
    elif step == "passwords":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        usernames = session["usernames"]
        if len(lines) != len(usernames):
            send_telegram_message(f"❌ পাসওয়ার্ড সংখ্যা ({len(lines)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। পুনরায় সঠিক লিস্ট দিন।", chat_id)
            return True
        session["passwords"] = lines
        session["step"] = "fa_keys"
        send_telegram_message("🔐 এখন **2FA কী** লিস্ট দিন (প্রতি লাইনে একটি, ইউজারনেম এর ক্রম অনুযায়ী):\n\nযদি 2FA না থাকে, লাইন ফাঁকা রাখবেন (শুধু এন্টার দিন)।", chat_id)
        return True
    elif step == "fa_keys":
        raw_lines = text.splitlines()
        fa_list = [l.strip() for l in raw_lines]
        usernames = session["usernames"]
        while len(fa_list) > len(usernames) and fa_list and fa_list[-1] == '':
            fa_list.pop()
        if len(fa_list) != len(usernames):
            send_telegram_message(f"❌ 2FA কী সংখ্যা ({len(fa_list)}) ইউজারনেম সংখ্যার ({len(usernames)}) সাথে মেলে না। প্রতিটি ইউজারনেমের জন্য একটি লাইন (খালিও হতে পারে) দিন।", chat_id)
            return True
        accounts_list = []
        for i in range(len(usernames)):
            accounts_list.append({"username": usernames[i], "password": session["passwords"][i], "fa_key": fa_list[i]})
        sell_id = uuid.uuid4().hex[:10]
        sell_req = {"id": sell_id, "user_id": chat_id, "accounts": accounts_list, "status": "pending", "time": time.time()}
        with data_lock:
            sell_requests.append(sell_req)
            save_sell_requests()
        save_data_to_channel()
        del sell_sessions[chat_id]
        sell_price = config.get("sell_price_per_account", 1.0)
        total_expected = sell_price * len(accounts_list)
        send_telegram_message(f"✅ আপনার বিক্রয় রিকোয়েস্ট জমা হয়েছে।\nআইডি: {sell_id}\nঅ্যাকাউন্ট সংখ্যা: {len(accounts_list)}\nপ্রত্যাশিত মূল্য: {total_expected} টাকা\nরিভিউ সম্পন্ন হতে ২৪ ঘণ্টা পর্যন্ত সময় লাগতে পারে।", chat_id)
        tg_username = user_info.get(str(chat_id), f"ID:{chat_id}")
        excel_bytes = generate_sell_excel(accounts_list, tg_username)
        filename = f"sell_{sell_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caption = f"📊 নতুন সেল রিকোয়েস্ট\nআইডি: {sell_id}\nইউজার: {tg_username} (`{chat_id}`)\nঅ্যাকাউন্ট সংখ্যা: {len(accounts_list)}\nপ্রত্যাশিত মূল্য: {total_expected} টাকা\nঅনুমোদন: /approvesell {sell_id} (অটো মূল্য {total_expected}) বা /approvesell {sell_id} <amount>"
        if not send_telegram_document(excel_bytes, filename, ADMIN_CHAT_ID, caption=caption):
            send_telegram_message(f"📊 সেল রিকোয়েস্ট (ফাইল পাঠানো যায়নি)\nআইডি: {sell_id}\nইউজার: {tg_username} ({chat_id})\nঅ্যাকাউন্ট: " + ", ".join([a['username'] for a in accounts_list]) + f"\nঅনুমোদন: /approvesell {sell_id} <amount>", ADMIN_CHAT_ID)
        send_main_keyboard(chat_id)
        return True
    return False

# ================== WITHDRAW ==================
def start_withdraw(chat_id):
    withdraw_sessions[chat_id] = {"step": "amount"}
    send_telegram_message("💸 উইথড্র\n\nআপনার কত টাকা উত্তোলন করতে চান? (শুধু সংখ্যা লিখুন)\n/start দিয়ে বাতিল করুন।", chat_id)

def process_withdraw_step(chat_id, text):
    if chat_id not in withdraw_sessions:
        return False
    if text.strip().lower() == "/start":
        del withdraw_sessions[chat_id]
        send_telegram_message("❌ উইথড্র বাতিল করা হয়েছে।", chat_id)
        send_main_keyboard(chat_id)
        return True
    session = withdraw_sessions[chat_id]
    step = session["step"]
    if step == "amount":
        try:
            amount = float(text)
            if amount <= 0:
                raise ValueError
        except:
            send_telegram_message("⚠️ সঠিক সংখ্যা দিন।", chat_id)
            return True
        with data_lock:
            bal = balances.get(str(chat_id), 0)
        if amount > bal:
            send_telegram_message(f"❌ পর্যাপ্ত ব্যালেন্স নেই। আপনার ব্যালেন্স: {bal} টাকা", chat_id)
            send_main_keyboard(chat_id)
            del withdraw_sessions[chat_id]
            return True
        session["amount"] = amount
        session["step"] = "bkash"
        send_telegram_message("📞 আপনার বিকাশ নম্বর দিন:", chat_id)
        return True
    elif step == "bkash":
        bkash = text.strip()
        if not bkash:
            send_telegram_message("⚠️ বিকাশ নম্বর দিন।", chat_id)
            return True
        amount = session["amount"]
        w_id = uuid.uuid4().hex[:10]
        w_req = {"id": w_id, "user_id": chat_id, "amount": amount, "bkash": bkash, "status": "pending", "time": time.time()}
        with data_lock:
            withdraw_requests.append(w_req)
            save_withdraw_requests()
        save_data_to_channel()
        del withdraw_sessions[chat_id]
        send_telegram_message(f"✅ আপনার {amount} টাকা উত্তোলনের রিকোয়েস্ট জমা হয়েছে।\nআইডি: {w_id}\nঅ্যাডমিন অ্যাপ্রুভ করলে টাকা পাঠানো হবে।", chat_id)
        admin_msg = f"💳 নতুন উইথড্র রিকোয়েস্ট\nআইডি: {w_id}\nইউজার: {user_info.get(chat_id, chat_id)} ({chat_id})\nপরিমাণ: {amount} টাকা\nবিকাশ: {bkash}\nঅনুমোদন: /approvewithdraw {w_id}\nবাতিল: /rejectwithdraw {w_id}"
        send_telegram_message(admin_msg, ADMIN_CHAT_ID)
        send_main_keyboard(chat_id)
        return True
    return False

# ================== ADMIN MARKETPLACE COMMANDS ==================
def admin_addstock_cmd(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    start_add_stock(chat_id)
    return True

def admin_stock_cmd(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        if not accounts:
            send_telegram_message("📭 কোনো অ্যাকাউন্ট স্টকে নেই।", chat_id)
        else:
            lines = [f"📦 স্টক ({len(accounts)} টি):"]
            for i, acc in enumerate(accounts, 1):
                lines.append(f"{i}. ইউজার: {acc['username']} | পাস: {acc['password']} | 2FA: {acc.get('fa_key', 'N/A')}")
            send_telegram_message("\n".join(lines), chat_id)
    return True

def admin_deletestock_cmd(chat_id, arg):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    try:
        idx = int(arg) - 1
        with data_lock:
            if 0 <= idx < len(accounts):
                deleted = accounts.pop(idx)
                save_accounts()
            else:
                send_telegram_message("❌ ভুল ইনডেক্স। /stock দিয়ে নম্বর দেখুন।", chat_id)
                return True
        save_data_to_channel()
        send_telegram_message(f"✅ স্টক থেকে অ্যাকাউন্ট `{deleted['username']}` মুছে ফেলা হয়েছে।", chat_id)
    except ValueError:
        with data_lock:
            for i, acc in enumerate(accounts):
                if acc["username"] == arg:
                    deleted = accounts.pop(i)
                    save_accounts()
                    save_data_to_channel()
                    send_telegram_message(f"✅ স্টক থেকে অ্যাকাউন্ট `{deleted['username']}` মুছে ফেলা হয়েছে।", chat_id)
                    return True
            send_telegram_message(f"❌ `{arg}` নামে কোনো অ্যাকাউন্ট পাওয়া যায়নি।", chat_id)
    return True

def admin_bulkdelete_cmd(chat_id, parts):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    if len(parts) < 3:
        send_telegram_message("❌ ফরম্যাট: /bulkdelete <শুরুর ইনডেক্স> <সংখ্যা> (যেমন /bulkdelete 1 5)", chat_id)
        return True
    try:
        start = int(parts[1]) - 1
        count = int(parts[2])
        if start < 0 or count <= 0:
            raise ValueError
    except:
        send_telegram_message("❌ সঠিক সংখ্যা দিন।", chat_id)
        return True
    with data_lock:
        if start + count > len(accounts):
            send_telegram_message(f"❌ স্টকে মোট {len(accounts)} টি অ্যাকাউন্ট আছে। আপনার নির্বাচিত রেঞ্জ অতিক্রম করছে।", chat_id)
            return True
        deleted = accounts[start:start+count]
        del accounts[start:start+count]
        save_accounts()
    save_data_to_channel()
    usernames = [acc['username'] for acc in deleted]
    send_telegram_message(f"✅ ইনডেক্স {start+1} থেকে {count} টি অ্যাকাউন্ট ডিলিট করা হয়েছে:\n" + "\n".join(usernames), chat_id)
    return True

def admin_setprice_cmd(chat_id, price_str):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    try:
        price = float(price_str)
        if price <= 0:
            raise ValueError
    except:
        send_telegram_message("❌ সঠিক মূল্য দিন (সংখ্যা)।", chat_id)
        return True
    with data_lock:
        config["price_per_account"] = price
        save_config()
    save_data_to_channel()
    send_telegram_message(f"✅ প্রতি অ্যাকাউন্টের মূল্য (ক্রয়) {price} টাকা নির্ধারণ করা হয়েছে।", chat_id)
    return True

def admin_setsellprice_cmd(chat_id, price_str):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    try:
        price = float(price_str)
        if price <= 0:
            raise ValueError
    except:
        send_telegram_message("❌ সঠিক মূল্য দিন (সংখ্যা)।", chat_id)
        return True
    with data_lock:
        config["sell_price_per_account"] = price
        save_config()
    save_data_to_channel()
    send_telegram_message(f"✅ বিক্রয়ের প্রতি অ্যাকাউন্টের মূল্য {price} টাকা নির্ধারণ করা হয়েছে।", chat_id)
    return True

def admin_setbkash_cmd(chat_id, number):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        config["bkash_number"] = number
        save_config()
    save_data_to_channel()
    send_telegram_message(f"✅ বিকাশ নম্বর {number} সেট করা হয়েছে।", chat_id)
    return True

def admin_setgroup_cmd(chat_id, group_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        config["group_chat_id"] = group_id
        save_config()
    save_data_to_channel()
    send_telegram_message(f"✅ ব্যাকআপ গ্রুপ আইডি {group_id} সেট করা হয়েছে।", chat_id)
    return True

def admin_setchannel_cmd(chat_id, channel_id_str):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    try:
        new_channel_id = int(channel_id_str)
        test_resp = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": new_channel_id, "text": "চ্যানেল কনফিগারেশন টেস্ট"}
        )
        if test_resp.status_code != 200 or not test_resp.json().get("ok"):
            send_telegram_message("❌ প্রদত্ত চ্যানেল আইডিতে মেসেজ পাঠানো যায়নি। নিশ্চিত করুন বট চ্যানেলের অ্যাডমিন।", chat_id)
            return True
        global CHANNEL_ID
        with data_lock:
            CHANNEL_ID = new_channel_id
            config["channel_id"] = str(CHANNEL_ID)
            save_config()
        save_data_to_channel()
        send_telegram_message(f"✅ ব্যাকআপ চ্যানেল {CHANNEL_ID} সেট করা হয়েছে।", chat_id)
    except ValueError:
        send_telegram_message("❌ সঠিক চ্যানেল আইডি সংখ্যা দিন।", chat_id)
    return True

def admin_approve_cmd(chat_id, deposit_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for dep in deposits:
            if dep["id"] == deposit_id and dep["status"] == "pending":
                dep["status"] = "approved"
                user = dep["user_id"]
                balances[str(user)] = balances.get(str(user), 0) + dep["amount"]
                save_balances()
                save_deposits()
                send_telegram_message(f"✅ ডিপোজিট {deposit_id} অনুমোদিত। ইউজারের ব্যালেন্স আপডেট হয়েছে।", chat_id)
                send_telegram_message(f"✅ আপনার {dep['amount']} টাকার ডিপোজিট অনুমোদিত হয়েছে। বর্তমান ব্যালেন্স: {balances[str(user)]} টাকা", user)
                break
        else:
            send_telegram_message("❌ ডিপোজিট পাওয়া যায়নি বা ইতিমধ্যে প্রসেস করা হয়েছে।", chat_id)
    save_data_to_channel()
    return True

def admin_reject_cmd(chat_id, deposit_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for dep in deposits:
            if dep["id"] == deposit_id and dep["status"] == "pending":
                dep["status"] = "rejected"
                save_deposits()
                send_telegram_message(f"❌ ডিপোজিট {deposit_id} বাতিল করা হয়েছে।", chat_id)
                send_telegram_message(f"❌ আপনার {dep['amount']} টাকার ডিপোজিট বাতিল করা হয়েছে।", dep["user_id"])
                break
        else:
            send_telegram_message("❌ ডিপোজিট পাওয়া যায়নি।", chat_id)
    save_data_to_channel()
    return True

def admin_deposits_cmd(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        pending = [d for d in deposits if d["status"] == "pending"]
    if not pending:
        send_telegram_message("কোনো পেন্ডিং ডিপোজিট নেই।", chat_id)
    else:
        lines = ["⏳ পেন্ডিং ডিপোজিট:"]
        for d in pending:
            lines.append(f"আইডি: {d['id']} | ইউজার: {d['user_id']} | পরিমাণ: {d['amount']} | ট্রানজেকশন: {d['trxid']}")
        send_telegram_message("\n".join(lines), chat_id)
    return True

def admin_addbalance_cmd(chat_id, uid, amt_str):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    try:
        amt = float(amt_str)
    except:
        send_telegram_message("❌ সঠিক পরিমাণ দিন।", chat_id)
        return True
    with data_lock:
        balances[str(uid)] = balances.get(str(uid), 0) + amt
        save_balances()
    save_data_to_channel()
    send_telegram_message(f"✅ {uid} এর ব্যালেন্সে {amt} টাকা যোগ করা হয়েছে। বর্তমান: {balances[str(uid)]}", chat_id)
    try:
        send_telegram_message(f"💰 অ্যাডমিন আপনার অ্যাকাউন্টে {amt} টাকা যোগ করেছেন। বর্তমান ব্যালেন্স: {balances[str(uid)]} টাকা", uid)
    except:
        pass
    return True

# ================== ADMIN SELL/WITHDRAW COMMANDS (UPDATED) ==================
def admin_approvesell_cmd(chat_id, sell_id, amount_str=None):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for req in sell_requests:
            if req["id"] == sell_id and req["status"] == "pending":
                if amount_str is None:
                    price = config.get("sell_price_per_account", 1.0)
                    amount = price * len(req["accounts"])
                else:
                    try:
                        amount = float(amount_str)
                    except:
                        send_telegram_message("❌ সঠিক পরিমাণ দিন।", chat_id)
                        return True
                req["status"] = "approved"
                user = req["user_id"]
                balances[str(user)] = balances.get(str(user), 0) + amount
                save_balances()
                save_sell_requests()
                send_telegram_message(f"✅ সেল রিকোয়েস্ট {sell_id} অনুমোদিত। {amount} টাকা ইউজারের ব্যালেন্সে যোগ হয়েছে।", chat_id)
                send_telegram_message(f"✅ আপনার বিক্রয় রিকোয়েস্ট অনুমোদিত হয়েছে। {amount} টাকা আপনার ব্যালেন্সে যোগ করা হয়েছে। বর্তমান ব্যালেন্স: {balances[str(user)]} টাকা", user)
                break
        else:
            send_telegram_message("❌ রিকোয়েস্ট পাওয়া যায়নি বা ইতিমধ্যে প্রসেস করা হয়েছে।", chat_id)
    save_data_to_channel()
    return True

def admin_rejectsell_cmd(chat_id, sell_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for req in sell_requests:
            if req["id"] == sell_id and req["status"] == "pending":
                req["status"] = "rejected"
                save_sell_requests()
                send_telegram_message(f"❌ সেল রিকোয়েস্ট {sell_id} বাতিল করা হয়েছে।", chat_id)
                send_telegram_message(f"❌ আপনার বিক্রয় রিকোয়েস্ট বাতিল করা হয়েছে।", req["user_id"])
                break
        else:
            send_telegram_message("❌ রিকোয়েস্ট পাওয়া যায়নি।", chat_id)
    save_data_to_channel()
    return True

def admin_approvewithdraw_cmd(chat_id, w_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for req in withdraw_requests:
            if req["id"] == w_id and req["status"] == "pending":
                user = req["user_id"]
                bal = balances.get(str(user), 0)
                if bal < req["amount"]:
                    send_telegram_message(f"❌ ইউজারের ব্যালেন্স {bal} টাকা, প্রয়োজন {req['amount']} টাকা।", chat_id)
                    return True
                req["status"] = "approved"
                balances[str(user)] = bal - req["amount"]
                save_balances()
                save_withdraw_requests()
                send_telegram_message(f"✅ উইথড্র {w_id} অনুমোদিত। ইউজারের {req['amount']} টাকা কেটে নেওয়া হয়েছে।", chat_id)
                send_telegram_message(f"✅ আপনার {req['amount']} টাকার উইথড্র অনুমোদিত হয়েছে। অ্যাডমিন শীঘ্রই বিকাশে টাকা পাঠাবেন।", user)
                break
        else:
            send_telegram_message("❌ রিকোয়েস্ট পাওয়া যায়নি।", chat_id)
    save_data_to_channel()
    return True

def admin_rejectwithdraw_cmd(chat_id, w_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        for req in withdraw_requests:
            if req["id"] == w_id and req["status"] == "pending":
                req["status"] = "rejected"
                save_withdraw_requests()
                send_telegram_message(f"❌ উইথড্র {w_id} বাতিল করা হয়েছে।", chat_id)
                send_telegram_message(f"❌ আপনার {req['amount']} টাকার উইথড্র বাতিল করা হয়েছে।", req["user_id"])
                break
        else:
            send_telegram_message("❌ রিকোয়েস্ট পাওয়া যায়নি।", chat_id)
    save_data_to_channel()
    return True

def admin_sell_requests_cmd(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        pending = [r for r in sell_requests if r["status"] == "pending"]
    if not pending:
        send_telegram_message("কোনো পেন্ডিং সেল রিকোয়েস্ট নেই।", chat_id)
    else:
        lines = ["📊 পেন্ডিং সেল রিকোয়েস্ট:"]
        for r in pending:
            acc_names = ", ".join([a["username"] for a in r["accounts"]])
            lines.append(f"আইডি: {r['id']} | ইউজার: {r['user_id']} | অ্যাকাউন্ট: {acc_names} | সংখ্যা: {len(r['accounts'])}")
            lines.append(f"অনুমোদন: /approvesell {r['id']} (অথবা /approvesell {r['id']} <amount>)")
            lines.append(f"বাতিল: /rejectsell {r['id']}\n")
        send_telegram_message("\n".join(lines), chat_id)
    return True

def admin_withdraw_requests_cmd(chat_id):
    if str(chat_id) != ADMIN_CHAT_ID:
        return False
    with data_lock:
        pending = [r for r in withdraw_requests if r["status"] == "pending"]
    if not pending:
        send_telegram_message("কোনো পেন্ডিং উইথড্র রিকোয়েস্ট নেই।", chat_id)
    else:
        lines = ["💳 পেন্ডিং উইথড্র রিকোয়েস্ট:"]
        for r in pending:
            lines.append(f"আইডি: {r['id']} | ইউজার: {r['user_id']} | পরিমাণ: {r['amount']} | বিকাশ: {r['bkash']}")
            lines.append(f"অনুমোদন: /approvewithdraw {r['id']}\nবাতিল: /rejectwithdraw {r['id']}\n")
        send_telegram_message("\n".join(lines), chat_id)
    return True

# ================== CENTRAL ADMIN DISPATCHER ==================
def handle_market_admin(chat_id, text):
    parts = text.split()
    if not parts:
        return False
    cmd = parts[0].lower()
    if cmd == "/addstock":
        return admin_addstock_cmd(chat_id)
    elif cmd == "/stock":
        return admin_stock_cmd(chat_id)
    elif cmd == "/deletestock":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /deletestock <ইনডেক্স অথবা ইউজারনেম>", chat_id)
            return True
        return admin_deletestock_cmd(chat_id, parts[1])
    elif cmd == "/bulkdelete":
        return admin_bulkdelete_cmd(chat_id, parts)
    elif cmd == "/setprice":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /setprice <মূল্য>", chat_id)
            return True
        return admin_setprice_cmd(chat_id, parts[1])
    elif cmd == "/setsellprice":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /setsellprice <মূল্য>", chat_id)
            return True
        return admin_setsellprice_cmd(chat_id, parts[1])
    elif cmd == "/setbkash":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /setbkash <বিকাশ নম্বর>", chat_id)
            return True
        return admin_setbkash_cmd(chat_id, parts[1])
    elif cmd == "/setgroup":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /setgroup <গ্রুপ চ্যাট আইডি>", chat_id)
            return True
        return admin_setgroup_cmd(chat_id, parts[1])
    elif cmd == "/setchannel":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /setchannel <চ্যানেল আইডি>", chat_id)
            return True
        return admin_setchannel_cmd(chat_id, parts[1])
    elif cmd == "/approve":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /approve <deposit_id>", chat_id)
            return True
        return admin_approve_cmd(chat_id, parts[1])
    elif cmd == "/reject":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /reject <deposit_id>", chat_id)
            return True
        return admin_reject_cmd(chat_id, parts[1])
    elif cmd == "/deposits":
        return admin_deposits_cmd(chat_id)
    elif cmd == "/addbalance":
        if len(parts) < 3:
            send_telegram_message("❌ ফরম্যাট: /addbalance <user_id> <amount>", chat_id)
            return True
        return admin_addbalance_cmd(chat_id, parts[1], parts[2])
    elif cmd == "/approvesell":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /approvesell <id> [amount]", chat_id)
            return True
        sell_id = parts[1]
        amount_str = parts[2] if len(parts) > 2 else None
        return admin_approvesell_cmd(chat_id, sell_id, amount_str)
    elif cmd == "/rejectsell":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /rejectsell <id>", chat_id)
            return True
        return admin_rejectsell_cmd(chat_id, parts[1])
    elif cmd == "/approvewithdraw":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /approvewithdraw <id>", chat_id)
            return True
        return admin_approvewithdraw_cmd(chat_id, parts[1])
    elif cmd == "/rejectwithdraw":
        if len(parts) < 2:
            send_telegram_message("❌ ফরম্যাট: /rejectwithdraw <id>", chat_id)
            return True
        return admin_rejectwithdraw_cmd(chat_id, parts[1])
    return False

# ================== MAIN COMMAND HANDLER ==================
def handle_telegram_commands():
    global last_update_id, maintenance_mode
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/getUpdates"
    while True:
        try:
            params = {"timeout": 30}
            if last_update_id:
                params["offset"] = last_update_id + 1
            response = requests.get(url, params=params, timeout=35)
            data = response.json()
            if data.get("ok") and data.get("result"):
                for update in data["result"]:
                    last_update_id = update["update_id"]
                    if "message" in update:
                        msg = update["message"]
                        chat_id = str(msg["chat"]["id"])
                        text = msg.get("text", "").strip()
                        from_user = msg.get("from", {})
                        sender_username = from_user.get("username") or from_user.get("first_name", f"ID:{chat_id}")
                        user_info[chat_id] = sender_username

                        if maintenance_mode and chat_id != ADMIN_CHAT_ID:
                            send_telegram_message("🔧 বট রক্ষণাবেক্ষণ মোডে আছে। পরে চেষ্টা করুন।", chat_id)
                            continue

                        if "document" in msg and chat_id in loss_recovery_sessions:
                            handle_loss_recovery_file(chat_id, msg)
                            continue

                        if "document" in msg and str(chat_id) == ADMIN_CHAT_ID:
                            caption = msg.get("caption", "").strip().lower()
                            if caption == "/restore":
                                file_id = msg["document"]["file_id"]
                                handle_restore(chat_id, file_id)
                            continue

                        if chat_id in support_sessions:
                            if text.lower() in ["/cancel", "/start"]:
                                support_sessions.discard(chat_id)
                                send_telegram_message("সাপোর্ট বাতিল।", chat_id)
                                send_main_keyboard(chat_id)
                            else:
                                forward = f"📩 সাপোর্ট মেসেজ\nইউজার: {sender_username} ({chat_id})\n\n{text}"
                                send_telegram_message(forward, ADMIN_CHAT_ID)
                                send_telegram_message("মেসেজ পাঠানো হয়েছে।", chat_id)
                                support_sessions.discard(chat_id)
                                send_main_keyboard(chat_id)
                            continue

                        if chat_id in deposit_sessions:
                            process_deposit_step(chat_id, text)
                            continue

                        if chat_id in submission_sessions:
                            process_submission_step(chat_id, text, sender_username)
                            continue

                        if chat_id in add_stock_sessions:
                            process_add_stock_step(chat_id, text)
                            continue

                        if chat_id in loss_recovery_sessions:
                            process_loss_recovery_step(chat_id, text)
                            continue

                        if chat_id in sell_sessions:
                            process_sell_step(chat_id, text)
                            continue

                        if chat_id in withdraw_sessions:
                            process_withdraw_step(chat_id, text)
                            continue

                        if chat_id in buy_sessions:
                            if text.strip().lower() in ["/cancel", "/start"]:
                                buy_sessions.discard(chat_id)
                                send_telegram_message("❌ কেনা বাতিল করা হয়েছে।", chat_id)
                                send_main_keyboard(chat_id)
                                continue
                            success = handle_buy(chat_id, text)
                            if success:
                                buy_sessions.discard(chat_id)
                                send_main_keyboard(chat_id)
                            continue

                        # --- Button Handlers ---
                        if text == "📋 সাবমিট":
                            start_submission(chat_id, sender_username)
                            continue
                        elif text == "🎁 মাদার একাউন্ট":
                            handle_getmother(chat_id)
                            continue
                        elif text == "📞 সাপোর্ট":
                            support_sessions.add(chat_id)
                            send_telegram_message("📞 আপনার সমস্যা বা প্রশ্ন লিখুন। অ্যাডমিন সরাসরি দেখতে পাবেন।\nবাতিল করতে /start বা /cancel লিখুন।", chat_id)
                            continue
                        elif text == "🛑 স্টপ":
                            with data_lock:
                                subscribed_users.discard(chat_id)
                                save_subscribers()
                            save_data_to_channel()
                            send_telegram_message("আপনার সাবস্ক্রিপশন বন্ধ করা হয়েছে।", chat_id, reply_markup=remove_keyboard())
                            continue
                        elif text == "💰 ব্যালেন্স":
                            bal = balances.get(chat_id, 0)
                            send_telegram_message(f"💰 আপনার বর্তমান ব্যালেন্স: {bal} টাকা", chat_id)
                            send_main_keyboard(chat_id)
                            continue
                        elif text == "💸 ডিপোজিট":
                            start_deposit(chat_id)
                            continue
                        elif text == "🛒 একাউন্ট কিনুন":
                            send_telegram_message("ইনস্টাগ্রাম কুকিজ এর কাজ যারা করেন আমাদের থেকে একাউন্ট কিনে কুকিজ সাবমিট দিলে যদি রিপোর্ট খারাপ হওয়ার কারণে আপনার লস হয় তবে আপনি লস রিকভারি অপশন থেকে আপনার যে পরিমাণ টাকা লস হবে তা ফেরত পাবেন।", chat_id)
                            buy_sessions.add(chat_id)
                            price = config.get("price_per_account", 1.70)
                            send_telegram_message(f"🛒 কতটি অ্যাকাউন্ট কিনতে চান? (সংখ্যা লিখুন, বাতিল করতে /start বা /cancel)\nপ্রতি অ্যাকাউন্টের মূল্য: {price} টাকা\nস্টক: {len(accounts)} টি", chat_id)
                            continue
                        elif text == "📥 ডিপোজিট রিকোয়েস্ট":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                admin_deposits_cmd(chat_id)
                            send_main_keyboard(chat_id)
                            continue
                        elif text == "➕ স্টক যোগ করুন":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                start_add_stock(chat_id)
                            continue
                        elif text == "📦 স্টক দেখুন":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                admin_stock_cmd(chat_id)
                            send_main_keyboard(chat_id)
                            continue
                        elif text == "🗑️ স্টক ডিলিট":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                send_telegram_message("🗑️ স্টক ডিলিট করতে কমান্ড ব্যবহার করুন:\n/deletestock <ইনডেক্স> বা /deletestock <ইউজারনেম>\nএকাধিক একসাথে ডিলিট: /bulkdelete <শুরুর ইনডেক্স> <সংখ্যা>\nস্টক দেখতে /stock দিন।", chat_id)
                            send_main_keyboard(chat_id)
                            continue
                        elif text == "🔄 লস রিকভারি":
                            start_loss_recovery(chat_id)
                            continue
                        elif text == "💰 বিক্রয় করুন":
                            start_sell(chat_id)
                            continue
                        elif text == "💸 উইথড্র":
                            start_withdraw(chat_id)
                            continue
                        elif text == "📊 সেল requests":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                admin_sell_requests_cmd(chat_id)
                            send_main_keyboard(chat_id)
                            continue
                        elif text == "💳 withdraw requests":
                            if str(chat_id) != ADMIN_CHAT_ID:
                                send_telegram_message("❌ অ্যাডমিন নন।", chat_id)
                            else:
                                admin_withdraw_requests_cmd(chat_id)
                            send_main_keyboard(chat_id)
                            continue

                        if text.startswith("/"):
                            if handle_market_admin(chat_id, text):
                                continue
                            if text.startswith("/start"):
                                with data_lock:
                                    subscribed_users.add(chat_id)
                                    save_subscribers()
                                save_data_to_channel()
                                send_telegram_message("✨ আমাদের বটে স্বাগতম! ✨", chat_id, reply_markup=get_keyboard(chat_id))
                                continue
                            elif text == "/stop":
                                with data_lock:
                                    subscribed_users.discard(chat_id)
                                    save_subscribers()
                                save_data_to_channel()
                                send_telegram_message("সাবস্ক্রিপশন বন্ধ করা হয়েছে।", chat_id, reply_markup=remove_keyboard())
                                continue
                            elif text.startswith("/addmother"):
                                args = text[len("/addmother"):].strip() if len(text) > len("/addmother") else ""
                                handle_addmother(chat_id, args)
                                continue
                            elif text == "/getmother":
                                handle_getmother(chat_id)
                                continue
                            elif text == "/motherlist":
                                handle_motherlist(chat_id)
                                continue
                            elif text.startswith("/deletemother"):
                                args = text[len("/deletemother"):].strip()
                                handle_deletemother(chat_id, args)
                                continue
                            elif text.startswith("/maintenance"):
                                args = text[len("/maintenance"):].strip()
                                handle_maintenance(chat_id, args)
                                continue
                            elif text == "/users":
                                handle_admin_users(chat_id)
                                continue
                            elif text.startswith("/broadcast"):
                                if len(text.split()) < 2:
                                    send_telegram_message("❌ ফরম্যাট: /broadcast <মেসেজ>", chat_id)
                                else:
                                    handle_admin_broadcast(chat_id, text.split(maxsplit=1)[1])
                                continue
                            elif text.startswith("/send"):
                                parts = text.split(maxsplit=2)
                                if len(parts) < 3:
                                    send_telegram_message("❌ ফরম্যাট: /send <user_id> <মেসেজ>", chat_id)
                                else:
                                    handle_admin_send(chat_id, parts[1], parts[2])
                                continue
                            elif text == "/backup":
                                handle_backup(chat_id)
                                continue
                            else:
                                send_telegram_message("❌ অজানা কমান্ড।", chat_id)
                                continue
        except Exception as e:
            logger.exception("Telegram Command Error:")
        time.sleep(1)

# ================== FLASK ROUTE ==================
@app.route("/")
def home():
    return "Bot Running Successfully!"

# ================== MAIN ==================
if __name__ == "__main__":
    load_mother_accounts()
    load_user_cooldowns()
    load_subscribers()
    load_market()
    load_sell_requests()
    load_withdraw_requests()

    auto_restore_from_channel()

    threading.Thread(target=handle_telegram_commands, daemon=True).start()
    threading.Thread(target=auto_backup_loop, daemon=True).start()

    save_data_to_channel()

    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
